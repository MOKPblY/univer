import requests
from openpyxl import Workbook
from univer.celery import app
from .report_settings import *

def get_ws_from_json(ws, title, headers, data):
    ws.title = title

    cols = len(headers)
    for col in range(1, cols+1):
        ws.cell(column=col, row=1, value=headers[col-1])

    for item, row in zip(data, range(2, len(data)+2)):
        for col, val in zip(range(1, cols + 1), item.values()):
            ws.cell(column=col, row=row, value=val)

    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(dims.get(cell.column_letter, 0), len(str(cell.value)))+5
    print(dims)
    for col, value in dims.items():
        ws.column_dimensions[col].width = value


@app.task(bind = True)
def get_and_save_report(self, filename='report.xlsx'):
    headers = {'Authorization': f'Api-Key {API_KEY}'}

    dirs_response = requests.get(url=DIRS_URL, headers=headers)
    dirs_data = dirs_response.json()

    groups_response = requests.get(url=GROUPS_URL, headers= headers)
    groups_data = groups_response.json()

    if dirs_response.status_code != 200 or groups_response.status_code != 200:
        raise Exception('Broker unauthorized')

    else:
        wb = Workbook()

        dirs_ws = wb.active
        get_ws_from_json(dirs_ws, 'Направления', DIRS_HEADERS, dirs_data)

        groups_ws = wb.create_sheet()
        get_ws_from_json(groups_ws, 'Группы', GROUPS_HEADERS, groups_data)

        wb.save(filename = filename)
        return True
